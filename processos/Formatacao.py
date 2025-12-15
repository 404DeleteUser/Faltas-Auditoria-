import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os

# Agora a função NÃO recebe argumentos, conforme seu pedido para o main.py
def executar_formatacao():
    
    # --- CONFIGURAÇÃO OBRIGATÓRIA (Caminho FIXO) ---
    
    # Adicionei o .xlsx no final, pois o openpyxl exige a extensão
    caminho_arquivo = r"C:\Users\08477936137\Downloads\AuditoriaDeFaltas\processos\Faltas\Auditoria_Completa_por_DRE.xlsx"

    # 2. Definição das cores (sem o '#')
    COR_CABECALHO = '356854'       # Verde escuro
    COR_TEXTO_CABECALHO = 'FFFFFF' # Branco
    COR_LINHA_BRANCA = 'FFFFFF'    # Branco
    COR_LINHA_ZEBRADA = 'beebda'   # Verde claro

    # --- FIM DA CONFIGURAÇÃO ---

    def auto_ajustar_colunas(ws):
        """Função auxiliar para ajustar a largura das colunas."""
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            try:
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            except Exception:
                pass

    def formatar_planilha_excel(caminho_interno):
        """Função interna que aplica a formatação."""
        caminho_str = str(caminho_interno)

        if not os.path.exists(caminho_str):
            print(f"ERRO: O arquivo não foi encontrado em: {caminho_str}")
            return

        print(f"Abrindo o arquivo: {caminho_str}...")
        
        try:
            wb = openpyxl.load_workbook(caminho_str)
            
            # Define Estilos
            fill_cabecalho = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
            font_cabecalho = Font(color=COR_TEXTO_CABECALHO, bold=True)
            fill_branca = PatternFill(start_color=COR_LINHA_BRANCA, end_color=COR_LINHA_BRANCA, fill_type="solid")
            fill_zebrada = PatternFill(start_color=COR_LINHA_ZEBRADA, end_color=COR_LINHA_ZEBRADA, fill_type="solid")

            print(f"Iniciando a formatação para {len(wb.worksheets)} aba(s)...")

            for ws in wb.worksheets:
                print(f"Formatando a aba: \"{ws.title}\"...", end='\r')
                
                max_linha = ws.max_row
                max_coluna = ws.max_column

                if max_linha == 0 or max_coluna == 0:
                    continue

                # 1. Cabeçalho
                for col_idx in range(1, max_coluna + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = fill_cabecalho
                    cell.font = font_cabecalho

                # 2. Zebrado
                for row_idx in range(2, max_linha + 1):
                    current_fill = fill_branca if row_idx % 2 == 0 else fill_zebrada
                    for col_idx in range(1, max_coluna + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = current_fill

                # 3. Congela Paineis
                ws.freeze_panes = 'A2'
                
                # 4. Ajuste Colunas
                auto_ajustar_colunas(ws)

            wb.save(caminho_str)
            print("\nFormatação concluída com sucesso!")

        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            if isinstance(e, PermissionError):
                print("ERRO: Feche o arquivo Excel antes de executar!")

    # Executa a função interna usando o caminho fixo definido no topo
    formatar_planilha_excel(caminho_arquivo)